"""Generate the synthetic messy test workbook used by the gates and tests.

Deliberate mess (Gate 1):
  - title block + merged banner cell, headers on row 3 (1-based)
  - a unit annotation row directly under the headers
  - a SUBTOTAL row inside the body and a TOTAL row at the bottom
  - the Logistics column is quoted in EUR ("EUR 0.12" strings)
  - messy header names ("P/N", "Qty/yr (EAU)", "Piece Price", "Weight (g)")

Planted data defects (Gate 2):
  - P-1010 has a NEGATIVE unit cost
  - P-1007's breakdown does not sum to unit_cost (>2% off)
  - P-1003 appears twice with CONFLICTING unit costs

Everything else is arithmetically consistent so reconciliation and the
analyses have clean ground truth. Base currency GBP.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

# part rows: (pn, name, commodity, supplier, region, eau, quote, unit_cost,
#             material, process, labour, overhead, sga, logistics_eur,
#             should_cost, target_cost, mass_g, grade)
ROWS = [
    ("P-1001", "Crankshaft forged", "Forging", "Alpha Forge", "EU", 120000, 42.50, 40.00, 22.00, 8.00, 4.00, 4.00, 2.00, 0.45, 36.10, 38.00, 14200, "42CrMo4"),
    ("P-1002", "Piston machined", "Machining", "Beta Machining", "EU", 480000, 11.80, 11.00, 4.40, 3.30, 1.65, 1.10, 0.55, 0.20, 10.20, 10.50, 350, "AlSi12"),
    ("P-1003", "Con rod", "Forging", "Alpha Forge", "EU", 240000, 9.90, 9.20, 4.60, 2.30, 1.15, 0.69, 0.46, 0.15, 8.75, 9.00, 620, "C70S6"),
    ("P-1004", "Oil pan stamped", "Sheet Metal", "Gamma Stamping", "Asia", 120000, 7.40, 7.00, 3.50, 1.75, 0.88, 0.52, 0.35, 0.30, 6.10, 6.50, 1850, "DC04"),
    ("P-1005", "Inverter housing", "Casting", "Delta Cast", "EU", 90000, 28.00, 26.50, 11.93, 7.95, 3.71, 1.86, 1.05, 0.60, 22.40, 24.00, 3400, "A380"),
    ("P-1006", "Stator housing", "Casting", "Delta Cast", "EU", 90000, 24.50, 23.00, 10.35, 6.90, 3.22, 1.61, 0.92, 0.55, 21.10, 22.00, 2950, "A380"),
    ("P-1007", "Battery tray", "Sheet Metal", "Gamma Stamping", "Asia", 60000, 55.00, 52.00, 20.00, 10.00, 5.00, 3.00, 2.00, 1.20, 44.30, 48.00, 12400, "AA6061"),  # breakdown sums to 40, unit 52 -> DEFECT
    ("P-1008", "Busbar PCBA", "PCBA", "Epsilon Elec", "Asia", 180000, 16.20, 15.50, 8.53, 3.88, 1.55, 1.09, 0.45, 0.25, 14.05, 14.80, 240, "FR4/Cu"),
    ("P-1009", "Fastener kit M8", "Fasteners", "Zeta Fast", "EU", 960000, 0.95, 0.88, 0.40, 0.22, 0.13, 0.09, 0.04, 0.02, 0.80, 0.85, 45, "8.8"),
    ("P-1010", "Bracket welded", "Sheet Metal", "Gamma Stamping", "EU", 30000, 4.20, -3.90, 1.75, 1.05, 0.55, 0.35, 0.20, 0.10, 3.55, 3.80, 900, "S355"),  # NEGATIVE unit cost -> DEFECT
    ("P-1003", "Con rod", "Forging", "Eta Forge", "Asia", 240000, 8.80, 8.40, 4.20, 2.10, 1.05, 0.63, 0.42, 0.12, 8.30, 8.60, 620, "C70S6"),  # duplicate PN, conflicting cost -> DEFECT
    ("P-1011", "Gear machined", "Machining", "Beta Machining", "EU", 200000, 14.90, 14.00, 5.60, 4.20, 2.10, 1.40, 0.70, 0.22, 12.60, 13.20, 780, "20MnCr5"),
    ("P-1012", "Housing machined", "Machining", "Theta CNC", "Asia", 150000, 21.00, 19.80, 7.92, 5.94, 2.97, 1.98, 0.99, 0.40, 17.90, 18.50, 2100, "AlSi9Cu3"),
    # Six more machining parts so cost-vs-mass regression clears n >= 8;
    # P-1018 sits well above the mass-cost line (planted VAVE candidate).
    ("P-1013", "Shaft machined", "Machining", "Beta Machining", "EU", 220000, 12.90, 12.20, 4.88, 3.66, 1.83, 1.22, 0.61, 0.18, 11.30, 11.80, 500, "42CrMo4"),
    ("P-1014", "Flange machined", "Machining", "Theta CNC", "Asia", 130000, 16.90, 16.00, 6.40, 4.80, 2.40, 1.60, 0.80, 0.24, 14.60, 15.20, 1200, "AlSi9Cu3"),
    ("P-1015", "Hub machined", "Machining", "Beta Machining", "EU", 110000, 19.00, 18.00, 7.20, 5.40, 2.70, 1.80, 0.90, 0.28, 16.40, 17.00, 1600, "20MnCr5"),
    ("P-1016", "Carrier machined", "Machining", "Theta CNC", "Asia", 95000, 25.40, 24.00, 9.60, 7.20, 3.60, 2.40, 1.20, 0.42, 21.80, 22.60, 2800, "AlSi9Cu3"),
    ("P-1017", "Diff case machined", "Machining", "Beta Machining", "EU", 80000, 29.10, 27.50, 11.00, 8.25, 4.13, 2.75, 1.37, 0.50, 25.10, 26.00, 3500, "GJS-600"),
    ("P-1018", "Valve body machined", "Machining", "Iota Precision", "EU", 140000, 22.80, 21.50, 8.60, 6.45, 3.23, 2.15, 1.07, 0.30, 15.90, 17.50, 1000, "AlSi9Cu3"),
]

HEADERS = [
    "P/N", "Description", "Commodity", "Vendor Name", "Region", "Qty/yr (EAU)",
    "Piece Price", "Unit Cost", "Material", "Process", "Direct Labour",
    "Overhead", "SG&A + Profit", "Logistics", "Should Cost", "Target Cost",
    "Weight (g)", "Material Grade",
]
UNIT_ROW = ["", "", "", "", "", "pcs", "£", "£", "£", "£", "£", "£", "£", "EUR", "£", "£", "g", ""]


def build(path: str | Path) -> Path:
    path = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Quote Summary"

    # Row 1: merged title banner. Row 2: blank. Row 3: headers. Row 4: units.
    ws.merge_cells("A1:R1")
    ws["A1"] = "PROJECT NEPTUNE — SUPPLIER QUOTATION SUMMARY (CONFIDENTIAL)"
    ws["A1"].font = Font(bold=True)

    ws.append([])  # row 2 blank
    ws.append(HEADERS)  # row 3
    ws.append(UNIT_ROW)  # row 4

    forging_rows = [r for r in ROWS if r[2] == "Forging"]
    other_rows = [r for r in ROWS if r[2] != "Forging"]

    def money(v: float) -> float:
        return v

    def append_part(r: tuple) -> None:
        (pn, name, comm, sup, reg, eau, quote, uc, mat, proc, lab, oh, sga, log_eur, sc, tc, mass_g, grade) = r
        ws.append([
            pn, name, comm, sup, reg, eau, money(quote), money(uc), money(mat), money(proc),
            money(lab), money(oh), money(sga), f"EUR {log_eur:.2f}", money(sc), money(tc), mass_g, grade,
        ])

    for r in forging_rows:
        append_part(r)
    # SUBTOTAL row inside the body (forging only).
    sub_uc = sum(r[7] for r in forging_rows)
    ws.append(["", "SUBTOTAL — Forging", "", "", "", "", "", sub_uc, "", "", "", "", "", "", "", "", "", ""])

    for r in other_rows:
        append_part(r)

    # TOTAL row at the bottom, arithmetically correct for Unit Cost.
    total_uc = sum(r[7] for r in ROWS)
    total_quote = sum(r[6] for r in ROWS)
    ws.append(["", "TOTAL", "", "", "", "", total_quote, total_uc, "", "", "", "", "", "", "", "", "", ""])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


if __name__ == "__main__":
    import sys

    out = sys.argv[1] if len(sys.argv) > 1 else "tests/data/synthetic_messy.xlsx"
    print(f"Wrote {build(out)}")
