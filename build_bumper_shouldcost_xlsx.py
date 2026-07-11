#!/usr/bin/env python3
"""
Tata Altroz Front Bumper Assembly — India should-cost workbook generator.

Produces Altroz-Front-Bumper-Should-Cost-India.xlsx with live formulas:
  Summary | Assumptions | Parts Costing | Bought-Out | Assembly & Packing | Tooling

All rates/masses/cycles are editable cells on Assumptions / Parts Costing;
every derived figure is an Excel formula, so the whole model recalculates.

Regenerate:  python3 build_bumper_shouldcost_xlsx.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

INDIGO = '4F46E5'
DARK   = '0F172A'
BODY   = '334155'
MUTED  = '64748B'
PANEL  = 'F1F5F9'
PANEL2 = 'EFF6FF'
GREEN  = '059669'
GREENBG= 'ECFDF5'
AMBER  = 'D97706'
AMBERBG= 'FFFBEB'
WHITE  = 'FFFFFF'

INR2 = '"₹" #,##0.00'
INR0 = '"₹" #,##0'
PCT  = '0.0%'
KG   = '0.000 "kg"'
SEC  = '0 "s"'

thin = Side(style='thin', color='CBD5E1')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()


def style_header_row(ws, row, cols, fill=INDIGO, color=WHITE, size=10):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color=color, size=size, name='Calibri')
        cell.fill = PatternFill('solid', fgColor=fill)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER


def title_block(ws, title, subtitle, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, size=16, color=WHITE, name='Calibri')
    c.fill = PatternFill('solid', fgColor=INDIGO)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c2 = ws.cell(row=2, column=1, value=subtitle)
    c2.font = Font(size=10, color=MUTED, italic=True, name='Calibri')
    c2.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[2].height = 18


# ═════════════════════════════ ASSUMPTIONS ═════════════════════════════
wa = wb.active
wa.title = 'Assumptions'
title_block(wa, 'CostVision — Should-Cost Assumptions & Rate Card (India, 2026)',
            'Every figure below is an input — edit any cell and the whole workbook recalculates. Sources: India tier-1 benchmarks, July 2026.', 4)

wa.column_dimensions['A'].width = 42
wa.column_dimensions['B'].width = 14
wa.column_dimensions['C'].width = 10
wa.column_dimensions['D'].width = 78


def a_section(row, label):
    wa.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = wa.cell(row=row, column=1, value=label)
    c.font = Font(bold=True, size=11, color=INDIGO, name='Calibri')
    c.fill = PatternFill('solid', fgColor=PANEL2)


def a_row(row, label, value, unit, note, fmt=None):
    wa.cell(row=row, column=1, value=label).font = Font(size=10, color=BODY, name='Calibri')
    v = wa.cell(row=row, column=2, value=value)
    v.font = Font(size=10, bold=True, color=DARK, name='Calibri')
    v.fill = PatternFill('solid', fgColor='FEF9C3')     # editable-input yellow
    v.border = BORDER
    if fmt: v.number_format = fmt
    wa.cell(row=row, column=3, value=unit).font = Font(size=9, color=MUTED, name='Calibri')
    wa.cell(row=row, column=4, value=note).font = Font(size=9, color=MUTED, name='Calibri')


a_section(3, 'PROGRAMME')
a_row(4,  'Annual vehicle volume', 72000, 'veh/yr', 'Altroz-class volume assumption — EDIT to your programme volume; all amortisation recalculates.', '#,##0')
a_row(5,  'Tooling amortisation period', 5, 'years', 'Tooling amortised into piece price over volume × years (standard should-cost practice).')
a_row(6,  'Currency', 'INR', '', 'All values in Indian Rupees, ex-works supplier (logistics to OEM excluded).')
a_row(7,  'FX reference', 106, 'INR/GBP', 'Reference only — no conversion applied in this workbook.', '#,##0')

a_section(9, 'MATERIAL PRICES  (₹/kg, delivered India plant)')
a_row(10, 'PP (black, textured grade)', 98, '₹/kg', 'Polypropylene copolymer, bulk OEM contract.', INR2)
a_row(11, 'PP+EPDM (TPO bumper grade)', 142, '₹/kg', 'Impact-modified bumper compound, paintable grade.', INR2)
a_row(12, 'ABS (plating/paint grade)', 148, '₹/kg', 'For fog-lamp bezels / DRL housings.', INR2)
a_row(13, 'EPP beads (45 g/l)', 330, '₹/kg', 'Expanded polypropylene for energy absorber.', INR2)
a_row(14, 'POM (clips)', 210, '₹/kg', 'Reference for clip costing (bought-out set used instead).', INR2)
a_row(15, 'High-strength steel sheet (HR, 590+)', 68, '₹/kg', 'For reinforcement beam & steel brackets.', INR2)
a_row(16, 'Steel scrap credit', 32, '₹/kg', 'Offal / engineered scrap sold back.', INR2)
a_row(17, 'Plastic regrind credit', 40, '₹/kg', 'Runner/sprue regrind value (where not reused).', INR2)
a_row(18, 'Paint material per bumper (primer+base+clear)', 160, '₹/unit', 'Body-colour robotic paint, material only.', INR2)

a_section(20, 'MACHINE RATES  (₹/hr — includes depreciation, energy @ ₹8.5/kWh, maintenance, floor space, supervision)')
a_row(21, 'Injection moulding 1500T', 2100, '₹/hr', 'For bumper fascia.', INR0)
a_row(22, 'Injection moulding 800T', 1150, '₹/hr', 'Spare class (unused by default).', INR0)
a_row(23, 'Injection moulding 450T', 700, '₹/hr', 'Grilles, air deflector.', INR0)
a_row(24, 'Injection moulding 250T', 520, '₹/hr', 'Bezels, wheel-arch retainers.', INR0)
a_row(25, 'Injection moulding 150T', 400, '₹/hr', 'Small brackets, holders, covers.', INR0)
a_row(26, 'EPP shape-moulding machine', 650, '₹/hr', 'Steam-chest moulding for energy absorber.', INR0)
a_row(27, 'Progressive press 250T (with feeder)', 900, '₹/hr', 'Steel beam brackets.', INR0)
a_row(28, 'Rollform + pierce + sweep line', 1600, '₹/hr', 'Reinforcement beam forming.', INR0)
a_row(29, 'MIG/spot weld cell', 800, '₹/hr', 'Beam bracket welding.', INR0)
a_row(30, 'Robotic paint line (bumper)', 2400, '₹/hr', 'Conveyorised, incl. booth energy & ventilation.', INR0)

a_section(32, 'LABOUR RATES  (₹/hr, fully loaded — wages + statutory + benefits + supervision share)')
a_row(33, 'Operator (semi-skilled)', 190, '₹/hr', 'Press/moulding/assembly operator.', INR0)
a_row(34, 'Skilled (setter / welder / painter)', 260, '₹/hr', '', INR0)
a_row(35, 'Quality inspector', 230, '₹/hr', 'End-of-line checks.', INR0)

a_section(37, 'COMMERCIAL FACTORS')
a_row(38, 'Factory overhead (on conversion cost)', 0.40, '%', 'Indirect staff, utilities, quality system, maintenance overheads.', PCT)
a_row(39, 'SG&A', 0.08, '%', 'Sales, general & administrative on manufacturing cost.', PCT)
a_row(40, 'Profit', 0.10, '%', 'Supplier margin on manufacturing cost + SG&A.', PCT)
a_row(41, 'Bought-out handling & inbound', 0.03, '%', 'Applied to bought-out (pass-through) items.', PCT)
a_row(42, 'Assembly overhead (on assembly conversion)', 0.45, '%', 'Line management, MHE, utilities at assembly plant.', PCT)
a_row(43, 'Packaging per finished assembly', 55, '₹/unit', 'Returnable dunnage trip cost, amortised.', INR2)
a_row(44, 'Parking sensors fitted', 4, 'pcs', 'Part list says 4–8 — set 4 (base) to 8 (top trim); sensors, holders & BOM recalc.', '0')

wa.freeze_panes = 'A3'

A = lambda cell: f'Assumptions!$B${cell}'   # shorthand


# ═════════════════════════════ PARTS COSTING ═════════════════════════════
wp = wb.create_sheet('Parts Costing')
HDRS = ['Sr', 'Part / Operation', 'Process', 'Material', 'Qty/veh',
        'Net mass (kg)', 'Scrap %', 'Mat. price (₹/kg)', 'Scrap credit (₹/kg)', 'Material ₹',
        'Machine / Line', 'Machine rate (₹/hr)', 'Cycle (s)', 'Cav.', 'Process ₹',
        'Manning', 'Labour rate (₹/hr)', 'Labour ₹',
        'Tool cost ₹ (total)', 'Tool amort ₹/pc', 'Overhead ₹', 'Mfg cost ₹',
        'SG&A ₹', 'Profit ₹', 'Unit price ₹', 'Extended ₹/veh']
NC = len(HDRS)
title_block(wp, 'Parts Costing — bottom-up build-up per manufactured part',
            'Yellow cells are inputs (masses/cycles are engineering estimates — replace with actuals when available). All ₹ columns are live formulas.', NC)
for i, h in enumerate(HDRS, 1):
    wp.cell(row=3, column=i, value=h)
style_header_row(wp, 3, range(1, NC + 1))
wp.row_dimensions[3].height = 30

# (name, process, material, qty(or formula), mass, scrap, matref, scrapref,
#  machine label, machine ref, cycle, cav, manning, labour ref, tool cost)
S = 'Q_SENS'  # marker: qty = sensor-count cell
parts = [
 ('Front bumper fascia', 'Injection moulding', 'PP+EPDM (TPO)', 1, 3.60, 0.04, 11, 17, 'IMM 1500T', 21, 55, 1, 0.5, 33, 18000000),
 ('Painting — fascia, body colour', 'Robotic paint (P+B+C)', 'Paint system', 1, None, None, None, None, 'Paint line', 30, 110, 1, 0.5, 34, 6000000),
 ('Upper grille', 'Injection moulding', 'PP (textured)', 1, 0.45, 0.05, 10, 17, 'IMM 450T', 23, 32, 1, 0.33, 33, 5000000),
 ('Lower grille', 'Injection moulding', 'PP (textured)', 1, 0.55, 0.05, 10, 17, 'IMM 450T', 23, 38, 1, 0.33, 33, 5000000),
 ('Side bracket LH+RH (fascia)', 'Injection moulding', 'PP', 2, 0.12, 0.04, 10, 17, 'IMM 150T', 25, 26, 2, 0.25, 33, 1800000),
 ('Number / licence plate bracket', 'Injection moulding', 'PP', 1, 0.18, 0.04, 10, 17, 'IMM 150T', 25, 28, 1, 0.25, 33, 1200000),
 ('Fog lamp bezel / DRL housing LH+RH', 'Injection moulding', 'ABS', 2, 0.15, 0.05, 12, 17, 'IMM 250T', 24, 34, 2, 0.33, 33, 2800000),
 ('Energy absorber', 'EPP shape moulding', 'EPP foam 45 g/l', 1, 0.25, 0.02, 13, 17, 'EPP moulder', 26, 90, 1, 0.5, 33, 3500000),
 ('Reinforcement beam — form & pierce', 'Rollform + sweep', 'HSS 590 sheet', 1, 4.20, 0.08, 15, 16, 'Rollform line', 28, 45, 1, 1.0, 33, 5000000),
 ('Reinforcement beam — bracket welding', 'MIG weld op', '(consumables in rate)', 1, None, None, None, None, 'Weld cell', 29, 60, 1, 1.0, 34, 1500000),
 ('Side beam bracket LH+RH (crash-box)', 'Progressive stamping', 'HSS 590 sheet', 2, 0.35, 0.12, 15, 16, 'Press 250T prog', 27, 12, 2, 0.5, 33, 2500000),
 ('Tow hook cover', 'Injection moulding', 'PP', 1, 0.05, 0.04, 10, 17, 'IMM 150T', 25, 22, 2, 0.25, 33, 900000),
 ('Parking sensor holder', 'Injection moulding', 'PP', S, 0.02, 0.04, 10, 17, 'IMM 150T', 25, 20, 4, 0.25, 33, 1400000),
 ('Radar sensor bracket', 'Injection moulding', 'PP', 1, 0.08, 0.04, 10, 17, 'IMM 150T', 25, 25, 2, 0.25, 33, 1100000),
 ('Camera mount bracket', 'Injection moulding', 'PP', 1, 0.05, 0.04, 10, 17, 'IMM 150T', 25, 22, 2, 0.25, 33, 900000),
 ('Air deflector', 'Injection moulding', 'PP', 1, 0.35, 0.05, 10, 17, 'IMM 450T', 23, 33, 1, 0.33, 33, 3000000),
 ('Wheel-arch retainer LH+RH', 'Injection moulding', 'PP', 2, 0.15, 0.04, 10, 17, 'IMM 250T', 24, 30, 2, 0.33, 33, 1600000),
 ('Side support bracket LH+RH', 'Injection moulding', 'PP', 2, 0.15, 0.04, 10, 17, 'IMM 250T', 24, 30, 2, 0.33, 33, 1600000),
]

r0 = 4
for idx, (name, proc, mat, qty, mass, scrap, matref, scrref, mach, machref,
          cyc, cav, man, labref, tool) in enumerate(parts):
    r = r0 + idx
    wp.cell(row=r, column=1, value=idx + 1)
    wp.cell(row=r, column=2, value=name)
    wp.cell(row=r, column=3, value=proc)
    wp.cell(row=r, column=4, value=mat)
    if qty == S:
        wp.cell(row=r, column=5, value=f'={A(44)}')
    else:
        wp.cell(row=r, column=5, value=qty)
    if mass is not None:
        wp.cell(row=r, column=6, value=mass).number_format = KG
        wp.cell(row=r, column=7, value=scrap).number_format = PCT
        wp.cell(row=r, column=8, value=f'={A(matref)}').number_format = INR2
        wp.cell(row=r, column=9, value=f'={A(scrref)}').number_format = INR2
        wp.cell(row=r, column=10, value=f'=F{r}*(1+G{r})*H{r}-F{r}*G{r}*I{r}').number_format = INR2
    else:
        # direct-material rows (paint / welding)
        if 'Paint' in proc or 'paint' in proc:
            wp.cell(row=r, column=10, value=f'={A(18)}').number_format = INR2
        else:
            wp.cell(row=r, column=10, value=0).number_format = INR2
    wp.cell(row=r, column=11, value=mach)
    wp.cell(row=r, column=12, value=f'={A(machref)}').number_format = INR0
    wp.cell(row=r, column=13, value=cyc).number_format = SEC
    wp.cell(row=r, column=14, value=cav)
    wp.cell(row=r, column=15, value=f'=M{r}*L{r}/3600/N{r}').number_format = INR2
    wp.cell(row=r, column=16, value=man)
    wp.cell(row=r, column=17, value=f'={A(labref)}').number_format = INR0
    wp.cell(row=r, column=18, value=f'=M{r}*P{r}*Q{r}/3600/N{r}').number_format = INR2
    wp.cell(row=r, column=19, value=tool).number_format = INR0
    wp.cell(row=r, column=20, value=f'=S{r}/({A(4)}*{A(5)}*E{r})').number_format = INR2
    wp.cell(row=r, column=21, value=f'=(O{r}+R{r})*{A(38)}').number_format = INR2
    wp.cell(row=r, column=22, value=f'=J{r}+O{r}+R{r}+T{r}+U{r}').number_format = INR2
    wp.cell(row=r, column=23, value=f'=V{r}*{A(39)}').number_format = INR2
    wp.cell(row=r, column=24, value=f'=(V{r}+W{r})*{A(40)}').number_format = INR2
    wp.cell(row=r, column=25, value=f'=V{r}+W{r}+X{r}').number_format = INR2
    wp.cell(row=r, column=26, value=f'=Y{r}*E{r}').number_format = INR2

rlast = r0 + len(parts) - 1
rt = rlast + 1
wp.cell(row=rt, column=2, value='MANUFACTURED PARTS — TOTAL PER VEHICLE')
wp.cell(row=rt, column=19, value=f'=SUM(S{r0}:S{rlast})').number_format = INR0
wp.cell(row=rt, column=26, value=f'=SUM(Z{r0}:Z{rlast})').number_format = INR2
for c in range(1, NC + 1):
    cell = wp.cell(row=rt, column=c)
    cell.fill = PatternFill('solid', fgColor=PANEL2)
    cell.font = Font(bold=True, size=10, color=DARK, name='Calibri')
    cell.border = BORDER

# styling pass
input_cols = {6, 7, 13, 14, 16, 19}          # engineering inputs (yellow)
for r in range(r0, rlast + 1):
    wp.row_dimensions[r].height = 24
    for c in range(1, NC + 1):
        cell = wp.cell(row=r, column=c)
        cell.border = BORDER
        if cell.font is None or not cell.font.bold:
            cell.font = Font(size=9.5, color=BODY, name='Calibri')
        cell.alignment = Alignment(vertical='center', wrap_text=(c == 2),
                                   horizontal=('left' if c in (2, 3, 4, 11) else 'center'))
        if c in input_cols and wp.cell(row=r, column=c).value is not None:
            cell.fill = PatternFill('solid', fgColor='FEF9C3')
        if r % 2 == 0 and c not in input_cols:
            cell.fill = PatternFill('solid', fgColor=PANEL)
        if c in (25, 26):
            cell.font = Font(size=9.5, bold=True, color=DARK, name='Calibri')

widths = [4, 34, 20, 17, 7, 10, 8, 11, 11, 11, 15, 11, 8, 6, 11, 8, 11, 10, 13, 10, 10, 11, 9, 9, 12, 12]
for i, w in enumerate(widths, 1):
    wp.column_dimensions[get_column_letter(i)].width = w
wp.freeze_panes = 'C4'


# ═════════════════════════════ BOUGHT-OUT ═════════════════════════════
wo = wb.create_sheet('Bought-Out')
title_block(wo, 'Bought-Out / Directed Parts — market reference prices (India OEM volumes)',
            'Electronics are typically customer-nominated buy items — prices are market references for completeness, not bottom-up should-costs.', 7)
BO_H = ['Sr', 'Item', 'Category', 'Qty/veh', 'Unit price ₹', 'Handling ₹', 'Landed ₹/veh']
for i, h in enumerate(BO_H, 1):
    wo.cell(row=3, column=i, value=h)
style_header_row(wo, 3, range(1, 8))

bo = [
 ('Ultrasonic parking sensor', 'Electronics', f'={A(44)}', 220),
 ('Front radar sensor (77 GHz)', 'Electronics', 1, 3200),
 ('Front camera module', 'Electronics', 1, 1500),
 ('Bumper wiring harness', 'Electrical', 1, 420),
 ('Clips & fasteners set (push clips, u-nuts, spring nuts)', 'Hardware', 1, 95),
 ('Bolts M8/M10 set', 'Hardware', 1, 28),
 ('Nuts & washers set', 'Hardware', 1, 20),
]
for idx, (item, catg, qty, price) in enumerate(bo):
    r = 4 + idx
    wo.cell(row=r, column=1, value=idx + 1)
    wo.cell(row=r, column=2, value=item)
    wo.cell(row=r, column=3, value=catg)
    wo.cell(row=r, column=4, value=qty)
    p = wo.cell(row=r, column=5, value=price)
    p.number_format = INR2
    p.fill = PatternFill('solid', fgColor='FEF9C3')
    wo.cell(row=r, column=6, value=f'=E{r}*{A(41)}').number_format = INR2
    wo.cell(row=r, column=7, value=f'=(E{r}+F{r})*D{r}').number_format = INR2
    for c in range(1, 8):
        cell = wo.cell(row=r, column=c)
        cell.border = BORDER
        if not cell.font.bold:
            cell.font = Font(size=10, color=BODY, name='Calibri')
        cell.alignment = Alignment(vertical='center', horizontal=('left' if c in (2, 3) else 'center'))
bo_last = 3 + len(bo)
rt = bo_last + 1
wo.cell(row=rt, column=2, value='BOUGHT-OUT — TOTAL PER VEHICLE').font = Font(bold=True, size=10, color=DARK, name='Calibri')
wo.cell(row=rt, column=7, value=f'=SUM(G4:G{bo_last})').number_format = INR2
rt2 = rt + 1
wo.cell(row=rt2, column=2, value='   of which electronics (sensors + radar + camera + harness)').font = Font(size=9.5, italic=True, color=MUTED, name='Calibri')
wo.cell(row=rt2, column=7, value=f'=SUM(G4:G7)').number_format = INR2
for c in range(1, 8):
    wo.cell(row=rt, column=c).fill = PatternFill('solid', fgColor=PANEL2)
    wo.cell(row=rt, column=c).font = Font(bold=True, size=10, color=DARK, name='Calibri')
    wo.cell(row=rt, column=c).border = BORDER
for i, w in enumerate([4, 48, 13, 9, 13, 11, 14], 1):
    wo.column_dimensions[get_column_letter(i)].width = w


# ═════════════════════════════ ASSEMBLY ═════════════════════════════
ws_ = wb.create_sheet('Assembly & Packing')
title_block(ws_, 'Assembly & Packing — station-by-station conversion cost',
            'Manned time per station × fully-loaded labour rate + fixture/equipment amortisation + overhead + returnable packaging.', 7)
AS_H = ['St.', 'Station / Operation', 'Time (s)', 'Operators', 'Labour rate (₹/hr)', 'Labour ₹/unit', 'Notes']
for i, h in enumerate(AS_H, 1):
    ws_.cell(row=3, column=i, value=h)
style_header_row(ws_, 3, range(1, 8))
stations = [
 ('Beam sub-assembly: energy absorber + crash brackets to beam', 60, 1.0, 33, 'Bolted, torque-monitored'),
 ('Fascia dress: grilles, bezels, number-plate bracket, tow cover', 120, 2.0, 33, 'Clip-fit + screw stations'),
 ('Harness routing, parking sensors, radar & camera fit', 150, 1.5, 33, 'Connector click-checks'),
 ('Air deflector, wheel-arch retainers, side supports', 75, 1.0, 33, ''),
 ('End-of-line inspection & torque verification', 60, 1.0, 35, 'Poka-yoke checklist'),
 ('Pack into returnable dunnage', 45, 1.0, 33, ''),
]
for idx, (name, t, ops, labref, note) in enumerate(stations):
    r = 4 + idx
    ws_.cell(row=r, column=1, value=idx + 1)
    ws_.cell(row=r, column=2, value=name)
    tc = ws_.cell(row=r, column=3, value=t); tc.number_format = SEC; tc.fill = PatternFill('solid', fgColor='FEF9C3')
    oc = ws_.cell(row=r, column=4, value=ops); oc.fill = PatternFill('solid', fgColor='FEF9C3')
    ws_.cell(row=r, column=5, value=f'={A(labref)}').number_format = INR0
    ws_.cell(row=r, column=6, value=f'=C{r}*D{r}*E{r}/3600').number_format = INR2
    ws_.cell(row=r, column=7, value=note)
    for c in range(1, 8):
        cell = ws_.cell(row=r, column=c)
        cell.border = BORDER
        cell.font = Font(size=10, color=BODY, name='Calibri')
        cell.alignment = Alignment(vertical='center', wrap_text=(c in (2, 7)),
                                   horizontal=('left' if c in (2, 7) else 'center'))
    ws_.row_dimensions[r].height = 24
st_last = 3 + len(stations)

extra = [
 ('Assembly labour subtotal', f'=SUM(F4:F{st_last})', ''),
 ('Fixtures & jigs amortisation (₹40 L over volume × years)', f'=4000000/({A(4)}*{A(5)})', 'Assembly fixtures, torque tools mounting'),
 ('Equipment amortisation (EOL tester, torque tools — ₹60 L)', f'=6000000/({A(4)}*{A(5)})', ''),
 ('Consumables per unit', 6, 'Tapes, primers, gloves'),
 ('Assembly overhead', None, 'On labour subtotal — rate from Assumptions'),
 ('Packaging (returnable dunnage trip cost)', f'={A(43)}', ''),
]
er0 = st_last + 2
for idx, (label, val, note) in enumerate(extra):
    r = er0 + idx
    ws_.cell(row=r, column=2, value=label).font = Font(size=10, bold=True, color=DARK, name='Calibri')
    if label == 'Assembly overhead':
        ws_.cell(row=r, column=6, value=f'=F{er0}*{A(42)}').number_format = INR2
    elif isinstance(val, str):
        ws_.cell(row=r, column=6, value=val).number_format = INR2
    else:
        c6 = ws_.cell(row=r, column=6, value=val); c6.number_format = INR2
        c6.fill = PatternFill('solid', fgColor='FEF9C3')
    ws_.cell(row=r, column=7, value=note).font = Font(size=9, color=MUTED, name='Calibri')
    for c in range(2, 8):
        ws_.cell(row=r, column=c).border = BORDER
tot_r = er0 + len(extra)
ws_.cell(row=tot_r, column=2, value='ASSEMBLY & PACKING — TOTAL PER VEHICLE')
ws_.cell(row=tot_r, column=6, value=f'=SUM(F{er0}:F{tot_r-1})').number_format = INR2
for c in range(2, 8):
    ws_.cell(row=tot_r, column=c).fill = PatternFill('solid', fgColor=PANEL2)
    ws_.cell(row=tot_r, column=c).font = Font(bold=True, size=10.5, color=DARK, name='Calibri')
    ws_.cell(row=tot_r, column=c).border = BORDER
for i, w in enumerate([5, 52, 9, 10, 13, 13, 34], 1):
    ws_.column_dimensions[get_column_letter(i)].width = w
ASSY_TOTAL = f"'Assembly & Packing'!$F${tot_r}"


# ═════════════════════════════ TOOLING ═════════════════════════════
wt = wb.create_sheet('Tooling')
title_block(wt, 'Tooling & Investment Summary',
            'One-time investment; amortised into piece price over volume × years (see Assumptions). Change any tool cost on Parts Costing.', 4)
wt.cell(row=3, column=1, value='Item').font = Font(bold=True, color=WHITE, name='Calibri')
wt.cell(row=3, column=2, value='Value').font = Font(bold=True, color=WHITE, name='Calibri')
style_header_row(wt, 3, range(1, 3))
trow = [
 ('Part tooling total (moulds, dies, fixtures)', f"='Parts Costing'!S{rlast+1}"),
 ('Assembly fixtures & jigs', 4000000),
 ('Assembly equipment (EOL, torque)', 6000000),
 ('TOTAL ONE-TIME INVESTMENT', None),
]
for idx, (label, val) in enumerate(trow):
    r = 4 + idx
    wt.cell(row=r, column=1, value=label)
    if val is None:
        wt.cell(row=r, column=2, value='=SUM(B4:B6)').number_format = INR0
        wt.cell(row=r, column=1).font = Font(bold=True, size=11, color=DARK, name='Calibri')
        wt.cell(row=r, column=2).font = Font(bold=True, size=11, color=DARK, name='Calibri')
        wt.cell(row=r, column=1).fill = PatternFill('solid', fgColor=PANEL2)
        wt.cell(row=r, column=2).fill = PatternFill('solid', fgColor=PANEL2)
    else:
        wt.cell(row=r, column=2, value=val).number_format = INR0
    for c in (1, 2):
        wt.cell(row=r, column=c).border = BORDER
wt.cell(row=9, column=1, value='In ₹ Crore').font = Font(size=10, italic=True, color=MUTED, name='Calibri')
wt.cell(row=9, column=2, value='=B7/10000000').number_format = '0.00 "Cr"'
wt.column_dimensions['A'].width = 46
wt.column_dimensions['B'].width = 18


# ═════════════════════════════ SUMMARY ═════════════════════════════
wsm = wb.create_sheet('Summary', 0)
title_block(wsm, 'TATA ALTROZ — FRONT BUMPER ASSEMBLY  ·  SHOULD-COST SUMMARY (INDIA)',
            'CostVision AI Cost Intelligence  ·  July 2026  ·  Currency: INR, ex-works  ·  Bottom-up physics build-up with live formulas', 6)

hdr = ['#', 'Cost block', 'Basis', '₹ / vehicle set', '% of total', '']
for i, h in enumerate(hdr, 1):
    wsm.cell(row=4, column=i, value=h)
style_header_row(wsm, 4, range(1, 6))

MFG = f"'Parts Costing'!$Z${rlast+1}"
BO  = f"'Bought-Out'!$G${bo_last+1}"
ELEC= f"'Bought-Out'!$G${bo_last+2}"
rows = [
 ('1', 'Manufactured parts (18 line items)', 'Bottom-up: material + process + labour + tooling amort + OH + margin', f'={MFG}'),
 ('2', 'Bought-out parts & hardware', 'Market reference + handling', f'={BO}'),
 ('3', 'Assembly & packing', 'Station labour + fixtures + overhead + dunnage', f'={ASSY_TOTAL}'),
]
for idx, (n, label, basis, formula) in enumerate(rows):
    r = 5 + idx
    wsm.cell(row=r, column=1, value=n)
    wsm.cell(row=r, column=2, value=label).font = Font(size=11, bold=True, color=DARK, name='Calibri')
    wsm.cell(row=r, column=3, value=basis).font = Font(size=9, color=MUTED, name='Calibri')
    wsm.cell(row=r, column=4, value=formula).number_format = INR2
    wsm.cell(row=r, column=5, value=f'=D{r}/$D$8').number_format = PCT
    for c in range(1, 6):
        wsm.cell(row=r, column=c).border = BORDER
        wsm.cell(row=r, column=c).alignment = Alignment(vertical='center', wrap_text=(c == 3))
    wsm.row_dimensions[r].height = 26
wsm.cell(row=8, column=2, value='TOTAL SHOULD COST — FRONT BUMPER ASSEMBLY').font = Font(bold=True, size=13, color=WHITE, name='Calibri')
wsm.cell(row=8, column=4, value='=SUM(D5:D7)').number_format = INR2
wsm.cell(row=8, column=4).font = Font(bold=True, size=13, color=WHITE, name='Calibri')
wsm.cell(row=8, column=5, value=1).number_format = PCT
wsm.cell(row=8, column=5).font = Font(bold=True, color=WHITE, name='Calibri')
for c in range(1, 6):
    wsm.cell(row=8, column=c).fill = PatternFill('solid', fgColor=INDIGO)
    wsm.cell(row=8, column=c).border = BORDER
wsm.row_dimensions[8].height = 30

wsm.cell(row=10, column=2, value='Mechanical module only (excl. electronics: sensors, radar, camera, harness)').font = Font(size=10.5, bold=True, color=BODY, name='Calibri')
wsm.cell(row=10, column=4, value=f'=D8-{ELEC}').number_format = INR2
wsm.cell(row=10, column=4).font = Font(bold=True, size=10.5, color=BODY, name='Calibri')
wsm.cell(row=11, column=2, value='One-time tooling & equipment investment (amortised in the prices above)').font = Font(size=10.5, color=MUTED, name='Calibri')
wsm.cell(row=11, column=4, value='=Tooling!B7').number_format = INR0
wsm.cell(row=12, column=2, value=f'Annual volume assumption').font = Font(size=10.5, color=MUTED, name='Calibri')
wsm.cell(row=12, column=4, value=f'={A(4)}').number_format = '#,##0 "veh/yr"'

wsm.cell(row=13, column=2, value='Painted fascia only (compare against fascia-only quotes)').font = Font(size=10.5, color=MUTED, name='Calibri')
wsm.cell(row=13, column=4, value="='Parts Costing'!Y4+'Parts Costing'!Y5").number_format = INR2

chart = BarChart()
chart.type = 'col'
chart.title = 'Should-cost breakdown (₹/vehicle set)'
chart.height = 8; chart.width = 16
data = Reference(wsm, min_col=4, min_row=5, max_row=7)
cats = Reference(wsm, min_col=2, min_row=5, max_row=7)
chart.add_data(data, titles_from_data=False)
chart.set_categories(cats)
chart.legend = None
chart.y_axis.title = '₹ per vehicle set'
wsm.add_chart(chart, 'A15')

notes_r = 32
wsm.cell(row=notes_r, column=1, value='KEY NOTES & BASIS').font = Font(bold=True, size=11, color=INDIGO, name='Calibri')
note_lines = [
 '1.  Basis: ex-works tier-1 supplier, India, 2026 rates. Edit any yellow cell (Assumptions / Parts Costing) — every figure recalculates.',
 '2.  Masses & cycle times are engineering estimates from typical B-segment bumper benchmarks — replace with actual CAD/part weights when available.',
 '3.  Electronics (radar, camera, sensors, harness) are customer-nominated buy items shown at market reference prices, not bottom-up should-costs.',
 '4.  Drawing notes PP for grilles while the part list says ABS — costed as PP (textured black, typical for this segment); switch material price ref on Parts Costing if ABS.',
 '5.  Parking sensor count set to 4 (base trim). Part list allows 4–8 — change one cell on Assumptions (B44) for higher trims.',
 '6.  Tooling amortised over volume × 5 years into piece price. For tooling paid separately by OEM, set tool costs to 0 on Parts Costing (column S).',
 '7.  Paint: fascia costed body-colour painted (material + robotic line). Delete row 5 on Parts Costing if supplied moulded-in-colour.',
 '8.  Logistics to OEM plant, warranty and ED/GST excluded (per should-cost convention).',
 '9.  COMPARING TO A QUOTE: match the scope first. A bumper supplier quote normally EXCLUDES radar/camera/sensors/harness (directed parts) and often the steel beam — '
 'compare it to the mechanical-module or painted-fascia lines above, not the headline total.',
]
for i, ln in enumerate(note_lines):
    wsm.cell(row=notes_r + 1 + i, column=1, value=ln).font = Font(size=9.5, color=BODY, name='Calibri')

for i, w in enumerate([6, 62, 46, 18, 11, 4], 1):
    wsm.column_dimensions[get_column_letter(i)].width = w

# ═════════════════════════════ METHODOLOGY & Q&A ═════════════════════════════
wm = wb.create_sheet('Methodology & Q&A', 1)
title_block(wm, 'How this should-cost was calculated — step by step',
            'Read top to bottom. The right-hand column shows LIVE numbers pulled by formula from the costing sheets (worked example: front bumper fascia).', 4)
wm.column_dimensions['A'].width = 7
wm.column_dimensions['B'].width = 30
wm.column_dimensions['C'].width = 100
wm.column_dimensions['D'].width = 20

def m_section(row, label):
    wm.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = wm.cell(row=row, column=1, value=label)
    c.font = Font(bold=True, size=12, color='FFFFFF', name='Calibri')
    c.fill = PatternFill('solid', fgColor=INDIGO)
    c.alignment = Alignment(vertical='center', indent=1)
    wm.row_dimensions[row].height = 22

def m_step(row, step, title, how, example_label=None, example_formula=None, fmt=INR2, height=52):
    wm.cell(row=row, column=1, value=step).font = Font(bold=True, size=14, color=INDIGO, name='Calibri')
    wm.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
    wm.cell(row=row, column=2, value=title).font = Font(bold=True, size=10.5, color=DARK, name='Calibri')
    wm.cell(row=row, column=2).alignment = Alignment(vertical='center', wrap_text=True)
    wm.cell(row=row, column=3, value=how).font = Font(size=10, color=BODY, name='Calibri')
    wm.cell(row=row, column=3).alignment = Alignment(vertical='center', wrap_text=True)
    if example_label:
        wm.cell(row=row, column=4, value=example_label).font = Font(size=8.5, color=MUTED, name='Calibri')
        wm.cell(row=row, column=4).alignment = Alignment(vertical='top', wrap_text=True, horizontal='center')
    if example_formula:
        v = wm.cell(row=row + 1, column=4, value=example_formula)
        v.number_format = fmt
        v.font = Font(bold=True, size=11, color=GREEN, name='Calibri')
        v.alignment = Alignment(horizontal='center', vertical='center')
        v.fill = PatternFill('solid', fgColor=GREENBG)
        v.border = BORDER
    for c in range(1, 5):
        wm.cell(row=row, column=c).border = BORDER
        if row % 2 == 0:
            pass
    wm.row_dimensions[row].height = height

m_section(3, 'THE 8-STEP METHOD  (worked example on the right: Front Bumper Fascia — values are LIVE, they change if you edit inputs)')
steps = [
 ('1', 'Break the assembly into parts and decide make vs buy',
  'The 26-line part list from the drawing was split into: 18 MANUFACTURED items (we cost these bottom-up on the Parts Costing sheet) and 7 BOUGHT-OUT items '
  '(electronics & standard hardware — no one should-costs a radar chip bottom-up; we use market reference prices + 3% handling on the Bought-Out sheet).',
  'Manufactured line items', None, INR2, 50),
 ('2', 'Material cost per part',
  'Net part mass × (1 + scrap%) × material price ₹/kg, minus scrap credit for the recoverable portion.  '
  'Formula in words: you buy slightly more plastic/steel than ends up in the part (runners, offal), and sell back what you can.  '
  'Excel: =Mass×(1+Scrap)×Price − Mass×Scrap×Credit  →  Fascia: 4.20 kg TPO at ₹142/kg with 4% runner scrap.',
  'Fascia material ₹', "='Parts Costing'!J4", INR2, 62),
 ('3', 'Process (machine) cost',
  'Cycle time ÷ 3600 × machine hour rate ÷ cavities.  The machine rate (Assumptions sheet) is a full build-up: depreciation + energy @ ₹8.5/kWh + maintenance '
  '+ floor space + supervision.  A 1500-tonne press at ₹2,100/hr running a 58 s cycle costs ₹33.83 of machine time per shot — big-ticket machines, small per-part cost, '
  'because the hour is shared across every part made in it.',
  'Fascia process ₹', "='Parts Costing'!O4", INR2, 62),
 ('4', 'Direct labour',
  'Cycle time × manning × labour rate ÷ 3600 ÷ cavities.  Manning below 1.0 means one operator tends several machines (standard moulding practice — 0.5 = one operator '
  'per two presses).  Labour rates are FULLY LOADED: wages + statutory contributions + benefits + supervision share.',
  'Fascia labour ₹', "='Parts Costing'!R4", INR2, 52),
 ('5', 'Tooling amortisation',
  'Tool cost ÷ (annual volume × amortisation years × parts per vehicle).  The ₹2.2 Cr fascia mould spread over 72,000 vehicles × 5 years = a per-part charge.  '
  'This is the number most sensitive to VOLUME — halve the volume and this line doubles.  If the OEM pays for tooling separately, set column S to zero on Parts Costing.',
  'Fascia tooling ₹/pc', "='Parts Costing'!T4", INR2, 55),
 ('6', 'Overheads and margin',
  'Factory overhead (40%) is applied on CONVERSION cost only (process + labour), never on material — a supplier does not "manage" resin price, so it earns no overhead on it.  '
  'Then SG&A 8% on manufacturing cost, and 10% profit on top of that.  The result is the ex-works unit price we should expect a competitive tier-1 to quote.',
  'Fascia unit price ₹', "='Parts Costing'!Y4", INR2, 55),
 ('7', 'Assembly & packing',
  'Six stations, each = manned seconds × operators × labour rate.  Add fixture and end-of-line equipment amortisation, 45% assembly overhead on the labour, consumables, '
  'and the returnable-dunnage trip cost.  Bumper assembly is clip-and-screw work — it is deliberately cheap (~1% of total).',
  'Assembly & packing ₹', f"={ASSY_TOTAL}", INR2, 52),
 ('8', 'Roll up the total',
  'Total = Σ(manufactured unit price × qty) + Σ(bought-out landed × qty) + assembly & packing.  The Summary sheet also shows the MECHANICAL-ONLY subtotal '
  '(excluding radar/camera/sensors/harness) because that is the number to compare against a bumper supplier\'s quote — the electronics are usually priced separately.',
  'TOTAL should cost ₹', "=Summary!D8", INR2, 55),
]
r = 4
for (n, t, how, exl, exf, fmt, h) in steps:
    m_step(r, n, t, how, exl, exf, fmt, h)
    r += 2 if exf else 1

qa_start = r + 1
m_section(qa_start, 'MANAGER Q&A — the questions you will get, and the answers (live numbers in the right column)')
qa = [
 ('“Why should I trust these numbers?”',
  'Because nothing is a black box. Every figure is built bottom-up from physics and visible inputs: mass × material price, cycle × machine rate, tool ÷ volume. '
  'Every input sits in a yellow cell you can challenge and change — and the whole workbook recalculates. A supplier quote can be compared against this line by line.',
  None, None),
 ('“Where do the rates come from?”',
  'India tier-1 benchmarks, July 2026: delivered polymer/steel contract prices, machine-hour build-ups at Indian energy (₹8.5/kWh) and wage levels, fully-loaded labour. '
  'They are on the Assumptions sheet with a note per line. If purchasing has better contract rates, type them in — that makes the model MORE accurate, not broken.',
  None, None),
 ('“₹13,000 for a bumper? The aftermarket part is ₹5,000!”',
  'Two different things. The ₹13.2k includes the ADAS electronics — radar, camera, 4 parking sensors, harness — which are ~79% of the total and are usually on separate '
  'commodity contracts. The mechanical bumper module (what a bumper supplier actually quotes) is the number on the right. Aftermarket MRP is also a retail price with '
  'distribution margins — not comparable to an OEM ex-works piece price.',
  'Mechanical module ₹', '=Summary!D10'),
 ('“What happens if the volume assumption is wrong?”',
  'One cell: Assumptions B4. Volume only touches amortisation (tooling, fixtures, equipment) — material, process and labour are per-piece and unaffected. '
  'At 72k/yr the fascia carries ₹61 of tooling; at 36k/yr it would carry ₹122. Test any scenario in seconds by editing the cell.',
  'Current volume', f'={A(4)}'),
 ('“Why is tooling inside the piece price?”',
  'Convention for should-cost comparisons — most Indian RFQs quote amortised piece price. If our programme pays tooling as a separate one-time invoice, zero out column S '
  'on Parts Costing; the piece prices drop and the Tooling sheet still shows the ₹ Cr investment to negotiate separately.',
  'Tooling investment ₹', '=Tooling!B7'),
 ('“Are the part weights real?”',
  'They are engineering estimates from typical B-segment bumper benchmarks (fascia 4.2 kg, beam 4.8 kg…). They sit in yellow cells on Parts Costing — replace them with '
  'CAD or weighed actuals and the model tightens immediately. Until then treat the result as a ±15–20% class estimate.',
  None, None),
 ('“The part list says ABS for the grilles but you costed PP?”',
  'The exploded drawing says PP, the part list says ABS — a genuine document conflict, flagged rather than hidden. PP (textured black) is typical for this segment, so PP '
  'was used. To switch: change the material price reference on the two grille rows to the ABS cell — roughly +₹50/kg on ~1.2 kg of parts.',
  None, None),
 ('“What is included and what is not?”',
  'Included: material, process, labour, tooling amortisation, factory overhead, SG&A, profit, assembly, returnable packaging — ex-works supplier. '
  'Excluded (by should-cost convention): freight to OEM plant, GST/duties, warranty provisions, ED&D. Add a logistics line if you need landed cost.',
  None, None),
 ('“Isn\'t 40% overhead too high (or too low)?”',
  'It is applied on conversion only (process + labour), NOT on material — so it is a much smaller ₹ than it sounds. 35–50% on conversion is the normal band for Indian '
  'tier-1 plastics/stamping plants. It is one editable cell (Assumptions B38) if our benchmark differs.',
  None, None),
 ('“How much margin did we allow the supplier?”',
  'SG&A 8% + profit 10%, both visible and editable. Together they are the negotiation band: a quote at our number ±5% is honest; a quote 25% above it needs a '
  'line-by-line conversation — and this workbook is exactly the agenda for that conversation.',
  None, None),
 ('“Which parts carry the most cost?”',
  'Painted fascia (~₹1,276 = moulding + body-colour paint), then the steel reinforcement beam (~₹495), then the grilles. Everything else is small brackets and covers '
  'below ₹120. In bought-out: the radar alone is more than the entire painted fascia.',
  'Painted fascia ₹', "='Parts Costing'!Y4+'Parts Costing'!Y5"),
 ('“Can we use this in a supplier negotiation?”',
  'Yes — that is its purpose. Hand the supplier the same structure and ask them to fill THEIR numbers: where their material price, cycle time or overhead differs from '
  'ours, that specific cell becomes the discussion. Suppliers concede far more against a bottom-up model than against “please give 5% discount”.',
  None, None),
 ('“Why does paint cost ₹416 when the paint material is only ₹185?”',
  'Body-colour painting is a process cost, not a material cost: a robotic paint line at ₹2,800/hr with booth energy and ventilation, ~130 s per bumper, plus overhead and '
  'margin on that conversion. If the fascia is supplied moulded-in-colour, delete the paint row — the model drops it cleanly.',
  'Paint line item ₹', "='Parts Costing'!Y5"),
 ('“4 parking sensors or 8?”',
  'Base trim = 4 (assumed), top trim = 8. One cell — Assumptions B44 — drives the sensor count, the holder count and the bought-out total together.',
  None, None),
 ('“How would we make this part cheaper?”',
  'The levers, in order of size: (1) delete/reduce paint (moulded-in-colour saves ~₹400), (2) resin contract price — every ₹10/kg on TPO is ~₹44 on the fascia, '
  '(3) volume pooling to spread tooling, (4) runner regrind reuse instead of selling scrap, (5) beam gauge/grade optimisation. The workbook quantifies each in seconds.',
  None, None),
]
r = qa_start + 1
for (q, a, exl, exf) in qa:
    wm.cell(row=r, column=1, value='Q').font = Font(bold=True, size=12, color=AMBER, name='Calibri')
    wm.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='center')
    wm.cell(row=r, column=2, value=q).font = Font(bold=True, size=10, color=DARK, name='Calibri')
    wm.cell(row=r, column=2).alignment = Alignment(vertical='center', wrap_text=True)
    wm.cell(row=r, column=3, value=a).font = Font(size=9.5, color=BODY, name='Calibri')
    wm.cell(row=r, column=3).alignment = Alignment(vertical='center', wrap_text=True)
    if exl:
        wm.cell(row=r, column=4, value=exl).font = Font(size=8.5, color=MUTED, name='Calibri')
        wm.cell(row=r, column=4).alignment = Alignment(vertical='top', wrap_text=True, horizontal='center')
        v = wm.cell(row=r + 1, column=4, value=exf)
        v.number_format = INR2 if 'volume' not in exl.lower() else '#,##0 "veh/yr"'
        v.font = Font(bold=True, size=11, color=GREEN, name='Calibri')
        v.alignment = Alignment(horizontal='center', vertical='center')
        v.fill = PatternFill('solid', fgColor=GREENBG)
        v.border = BORDER
    for c in range(1, 5):
        wm.cell(row=r, column=c).border = BORDER
        if (r - qa_start) % 2 == 0:
            if not wm.cell(row=r, column=c).fill.fgColor.rgb or wm.cell(row=r, column=c).fill.fgColor.rgb == '00000000':
                wm.cell(row=r, column=c).fill = PatternFill('solid', fgColor=PANEL)
    wm.row_dimensions[r].height = 66
    r += 2 if exl else 1
wm.freeze_panes = 'A3'

OUT = 'Altroz-Front-Bumper-Should-Cost-India.xlsx'
wb.save(OUT)
print(f'Wrote {OUT} with sheets: {wb.sheetnames}')
